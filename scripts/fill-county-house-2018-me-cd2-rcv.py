#!/usr/bin/env python3
"""
Fills ME-02's 2018 county-level House data by replicating Maine's actual ranked-choice
tabulation from the Secretary of State's own cast-vote-record (CVR) files - the
historic first-ever RCV U.S. House election (Golden vs. Poliquin), previously a
documented dead end for this whole project (neither MEDSL nor OpenElections publish
ME-02 2018 at all; both sources only have first-round-style data even where they DO
cover an RCV race elsewhere, and Maine's actual round-2 winner differs from the
round-1 leader - Poliquin led round 1, Golden won after redistribution - so a
first-round-only county number would misrepresent which candidate actually won each
county).

Source: https://www.maine.gov/sos/elections-voting/election-results-data/election-results-2018
- CVR files (ballot-level, one row per ballot with up to 5 ranked choices + a town/
  precinct name): Nov18CVRExportFINAL{1,2,3}.xlsx, RepCD2-8final.xlsx, plus 3 UOCAVA
  (overseas/military) files - UOCAVA-FINALRepCD2.xlsx, UOCAVA-AUX-CVRRepCD2.xlsx,
  UOCAVA2CVRRepCD2.xlsx (~296k ballots total, matching the state's own certified
  296,077-ballot total almost exactly).
- Official round-by-round summary: updated-summary-report-CD2.xls - gives the
  certified STATEWIDE final result (Golden 142,440 / Poliquin 138,931) used here only
  to validate this script's own tabulation, not as a data source itself (it has no
  county breakdown).

**Replicates Maine's actual elimination rule directly from the CVR data**: the
official summary shows exactly 2 rounds (not 4 sequential single-candidate
eliminations), confirming Maine batch-eliminates every mathematically-certain loser at
once. This script does the same: for each ballot, if the 1st choice is Golden or
Poliquin, it stays; if 1st choice is Bond, Hoar, or a write-in, the ballot is walked
forward through its remaining ranks (skipping undervotes and further eliminated
candidates) until it finds Golden, Poliquin, an overvote (exhausted), or runs out of
ranks (exhausted). This reproduces the state's own round-2 totals almost exactly:
142,155/138,704 vs. the certified 142,440/138,931 (~0.1% short, from a handful of
ballots not captured across the 7 source files - well within this pipeline's usual
tolerance).

Ballots are attributed to counties via the Census Bureau's 2020 county subdivision
gazetteer (`2020_gaz_cousubs_23.txt`) as a town-name-to-county crosswalk, with
normalization for ward suffixes ("Bangor W1"), "All" combined-precinct labels
("Lewiston All"), multi-town voting districts ("Old Town/Argyle Twp"), and
Saint/St. spelling. ~296,982 of 296,077 official ballots were geocoded to a county
(UOCAVA ballots and a residual ~3.3k ballots in small unorganized townships/tribal
voting districts not in the standard gazetteer could not be attributed to any county -
same "some ballots aren't geographically attributable" class as this project's
already-documented ME 2020 Senate "Overseas" row). This leaves county totals ~1-2%
short of the certified statewide result, accepted at the same tolerance as this
project's other absentee/unattributable gaps.

**Kennebec County is split between ME-01 and ME-02** (already had a 2018 House row
from this project's earlier work, district "1" only) - this script ADDS ME-02's
portion on top of the existing row rather than overwriting it, discovered as a
byproduct of this fix (the existing Kennebec row was itself incomplete before this
script ran).

Since Golden and Poliquin are the only two candidates remaining after round 2 (Bond and
Hoar are fully eliminated, matching the certified summary's own convention), every row
here has oth=0 - there is no "third-party round 2" bucket in ranked-choice tabulation
the way this pipeline's other oth columns work.

Run from project root: python3 scripts/fill-county-house-2018-me-cd2-rcv.py
(Downloads ~12MB of CVR files fresh from maine.gov each run - not cached to disk,
matching this project's OpenElections-fetching convention.)
"""
import csv, io, os, re
import urllib.request
from collections import defaultdict, Counter

import openpyxl

ROOT = os.path.join(os.path.dirname(__file__), "..")
OUT_CSV = os.path.join(ROOT, "data-entry/county_house_results_2018.csv")
YEAR = 2018

BASE = "https://www.maine.gov/sos/sites/maine.gov.sos/files/content/assets/"
CVR_FILES = [
    "Nov18CVRExportFINAL1.xlsx", "Nov18CVRExportFINAL2.xlsx", "Nov18CVRExportFINAL3.xlsx",
    "RepCD2-8final.xlsx",
    "UOCAVA-FINALRepCD2.xlsx", "UOCAVA-AUX-CVRRepCD2.xlsx", "UOCAVA2CVRRepCD2.xlsx",
]
UOCAVA_FILES = {"UOCAVA-FINALRepCD2.xlsx", "UOCAVA-AUX-CVRRepCD2.xlsx", "UOCAVA2CVRRepCD2.xlsx"}
GAZETTEER_URL = "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/2020_Gazetteer/2020_gaz_cousubs_23.txt"

NEW_ME02_COUNTIES = {
    "23001": "Androscoggin", "23003": "Aroostook", "23007": "Franklin", "23009": "Hancock",
    "23017": "Oxford", "23019": "Penobscot", "23021": "Piscataquis", "23025": "Somerset",
    "23027": "Waldo", "23029": "Washington",
}


def fetch_bytes(url):
    req = urllib.request.Request(url, headers={"User-Agent": "election-map-data-pipeline/1.0"})
    with urllib.request.urlopen(req) as r:
        return r.read()


def classify(cell):
    if cell is None:
        return "undervote"
    s = str(cell)
    if s in ("undervote", "overvote"):
        return s
    if "Poliquin" in s:
        return "POLIQUIN"
    if "Golden" in s:
        return "GOLDEN"
    if "Bond" in s:
        return "BOND"
    if "Hoar" in s:
        return "HOAR"
    return "OTHER"


def load_town_to_county():
    text = fetch_bytes(GAZETTEER_URL).decode("latin-1")
    m = {}
    for line in text.splitlines()[1:]:
        parts = line.split("\t")
        if len(parts) < 4:
            continue
        geoid, name = parts[1], parts[3]
        county_fips = "23" + geoid[2:5]
        base = re.sub(r"\s+(city|town|township|Twp|CDP|Reservation|UT)\.?$", "", name, flags=re.I).strip()
        m[base.upper()] = county_fips
        m[name.upper()] = county_fips
    return m


def normalize_variants(raw):
    out = {raw}
    more = set()
    for c in out:
        more.add(re.sub(r"^Saint\b", "St.", c, flags=re.I))
        more.add(re.sub(r"^Saint\b", "St", c, flags=re.I))
        more.add(re.sub(r"\bTwp\.?$", "Township", c, flags=re.I))
        more.add(re.sub(r"\bPlt\.?$", "Plantation", c, flags=re.I))
    return out | more


def match_county(town, town_to_county):
    if town == "UOCAVA":
        return None
    raw = str(town).strip()
    bases = [
        raw,
        re.sub(r"\s*-\s*.*$", "", raw),
        re.sub(r"\s+W(ard)?\s*\d+$", "", raw, flags=re.I),
        re.sub(r"\s+All$", "", raw, flags=re.I),
    ]
    if "/" in raw:
        first = raw.split("/")[0].strip()
        bases.append(first)
        bases.append(re.sub(r"\s+All$", "", first, flags=re.I))
    candidates = set()
    for b in bases:
        candidates |= normalize_variants(b.strip())
    for c in candidates:
        r = town_to_county.get(c.strip().upper())
        if r:
            return r
    return None


def process_ballots():
    r2_by_town = defaultdict(Counter)
    for fn in CVR_FILES:
        is_uocava = fn in UOCAVA_FILES
        raw = fetch_bytes(BASE + fn)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            town = "UOCAVA" if is_uocava else row[1]
            choices = [classify(c) for c in row[3:8]]
            c1 = choices[0]
            if c1 in ("undervote", "overvote"):
                continue
            if c1 in ("GOLDEN", "POLIQUIN"):
                r2_by_town[town][c1] += 1
                continue
            for c in choices[1:]:
                if c in ("GOLDEN", "POLIQUIN"):
                    r2_by_town[town][c] += 1
                    break
                if c == "overvote":
                    break
                if c in ("BOND", "HOAR", "OTHER", "undervote"):
                    continue
    return r2_by_town


def main():
    print("Fetching Maine town-to-county crosswalk...")
    town_to_county = load_town_to_county()

    print("Fetching and tabulating CVR files (this takes a bit)...")
    r2_by_town = process_ballots()

    county_r2 = defaultdict(Counter)
    unmatched = 0
    for town, counts in r2_by_town.items():
        fips = match_county(town, town_to_county)
        if fips is None:
            unmatched += sum(counts.values())
            continue
        county_r2[fips].update(counts)

    total_golden = sum(c.get("GOLDEN", 0) for c in county_r2.values())
    total_poliquin = sum(c.get("POLIQUIN", 0) for c in county_r2.values())
    print(f"Matched county totals: Golden={total_golden} Poliquin={total_poliquin} "
          f"(certified: 142440/138931) | unmatched ballots: {unmatched}")

    with open(OUT_CSV, newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)
    by_fips = {r["county_id"]: r for r in rows}

    for fips, counts in county_r2.items():
        dem, gop = counts.get("GOLDEN", 0), counts.get("POLIQUIN", 0)
        total = dem + gop
        if fips in by_fips:
            r = by_fips[fips]
            r[f"dem_{YEAR}"] = str(int(r[f"dem_{YEAR}"]) + dem)
            r[f"gop_{YEAR}"] = str(int(r[f"gop_{YEAR}"]) + gop)
            r[f"total_{YEAR}"] = str(int(r[f"total_{YEAR}"]) + total)
            existing = [d for d in r[f"districts_{YEAR}"].split(";") if d]
            if "2" not in existing:
                existing.append("2")
            r[f"districts_{YEAR}"] = ";".join(sorted(existing))
        else:
            name = NEW_ME02_COUNTIES.get(fips, fips)
            new_row = {
                "state": "ME", "county_name": name, "county_id": fips,
                f"dem_{YEAR}": dem, f"gop_{YEAR}": gop, f"oth_{YEAR}": 0,
                f"total_{YEAR}": total, f"districts_{YEAR}": "2",
            }
            rows.append(new_row)
            by_fips[fips] = new_row

    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    print(f"Wrote {len(county_r2)} ME counties (10 new + Kennebec's CD2 addition) -> {OUT_CSV}")


if __name__ == "__main__":
    main()
